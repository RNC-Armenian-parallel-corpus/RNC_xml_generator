import hy_translit
import razdel
from openpyxl import load_workbook
import os
import re
from uniparser_eastern_armenian import EasternArmenianAnalyzer
import lxml.etree as ET

import eanc2rnc_tag_converter as tag_converter

import time
import functools

analyzer = EasternArmenianAnalyzer()
print('Analyzer imported')

def timer(func):
    @functools.wraps(func)
    def _timer(*args, **kwargs):
        start = time.perf_counter()
        result = func(*args, **kwargs)
        end = time.perf_counter()
        print(f"Execution time: {end - start:.4f} seconds")
        return result

    return _timer

AM_PUNCT_CORRECTIONS = str.maketrans({':': '։', '`': '՝'}) # заменяем знаки пунктуации

class XLSX2XML:

    def __init__(self, filename, col_mapping, input_path, output_path, ru_if_annotate = False):
        self.FILENAME = filename
        self.INPUT_PATH = input_path
        self.OUTPUT_PATH = output_path
        self.FILEPATH = os.path.join(input_path, filename)

        self.LANG_COLUMNS = col_mapping
        self.am_tokenize = razdel.tokenize
        # self.ru_tokenize = razdel.tokenize
        self.ru_if_annotate = ru_if_annotate

        self.header = self.load_aligned()

        self.am_word_count = 0 # количество слов в тексте на армянском
        self.ru_word_count = 0 # количество слов в тексте на русском
        self.stats = () # словарь для записи количества предложений и слов в файле (напр., ('news1.xlsx', 6, 83) )

    def load_aligned(self):
        wb = load_workbook(self.FILEPATH, read_only=True)
        self.ws = wb.active
        for row in self.ws.iter_rows(max_row=1, values_only=True):
            header = list(filter(lambda s: s in self.LANG_COLUMNS.keys(), row))
            assert len(header)==2, 'Не получилось найти перечисленные столбцы!'
            print('Used columns:', header)
            return header

    @staticmethod
    def convolve_whitespace(text):
        return re.sub(r'[\s(\xa0)]+', ' ', text)

    @staticmethod
    def translit(word):
        return hy_translit.transliterate_MEA(word)

    def annotate_ru(self, se, sent, if_annotate=False):
        self.ru_word_count += len(re.findall(r'\s', sent)) + 1
        if re.search(r'\w', sent):
            se.set('lang', 'rus')
            if not if_annotate:
                se.text = sent
                return se

            # tagged = []

            # ru_sent_analysis = mystem.analyze(unescape(sent))
            # for w in ru_sent_analysis:
            #     new_ana = adel.process_parse(w)
            #     tagged.append(new_ana)
            # return ''.join(tagged)  # has not been tested

    def annotate_am(self, se, sent):
        # correct some OCR in Armenian text
        sent = sent.translate(AM_PUNCT_CORRECTIONS)

        self.am_word_count += len(re.findall(r'\s+', sent)) + 1

        if re.search(r'\w', sent):
            se.text = ''
            se.set('lang', 'hye')
            se.set('has_translit', 'true')

            last_word = None

            tokens = [t.text for t in self.am_tokenize(sent)]

            for t in tokens:
                # если в токене есть буквы или цифры
                if re.search(r'\w', t):

                    w = ET.SubElement(se, "w")
                    tr = self.translit(t)
                    if tr:
                        w.set('translit', tr)

                    analysis = list(filter(lambda a: a['lemma'] or a['gramm'], analyzer.analyze_words(words=t, format='json')))
                    # если есть разбор и он не для числа
                    if analysis and not re.search(r'[0-9]', t):
                        for a in analysis:

                            # заменяем теги в грамматическом разборе на те, что используются в НКРЯ
                            gram_ana = tag_converter.convert_tags(','.join(a.get('gramm', ''))) # получается, что теги сначала слепляются через запятую, а потом разлепляются функцией, что suboptimal, но это проблемы нас в будущем

                            ana = ET.SubElement(w, "ana",)
                            ana.set('lex', a.get('lemma', ''))
                            ana.set('gr', gram_ana.replace(',',' '))
                            ana.set('transl', a.get('trans_en',''))
                            ana.set('lex_translit', self.translit(a.get('lemma','')))

                        ana.tail = t
                    # если анализа нет или токен -- число
                    else:
                        # w = ET.SubElement(se, "w")
                        # w.set('translit', self.translit(t))
                        w.text = t

                    last_word = w
                    last_word.tail = ''
                # если это пробел или знак препинания, то вставляем токен после последнего w
                else:
                    try:
                        last_word.tail += t
                    except AttributeError:
                        # если предложение начинается со знака препинания, то не существует w,
                        # после которого можно вставить токен. Поэтому вставляем его как текст
                        se.text += t

            # проверяем, что тексты до и после обработки сходятся
            res_sent = ET.tostring(se, method="text", encoding='unicode')

            if sent != res_sent:
                print('The resulting sentence does not match the original!')
                print('original:', sent, sep='\t')
                print('result:', res_sent, sep='\t\t')
                print('')

            return se

    def insert_into_tree(self, para_, sent_, lang):
        sent = self.convolve_whitespace(sent_)
        if not sent or sent==' ':
            return None
        se = ET.SubElement(para_, "se")
        if lang == 'rus':
            self.annotate_ru(se, sent)
        elif lang == 'hye':
            self.annotate_am(se, sent)
        return para_

    @timer
    def write_xml(self, move_processed=False):

        result_xml = os.path.join(self.OUTPUT_PATH, self.FILENAME.split('.')[-2] + '_processed.xml')

        with open(result_xml, 'w', encoding="utf-8") as f:
            f.write('<?xml version="1.0" encoding="utf-8"?>\n')
            f.write('<root>\n')
            f.write('<head>\n')
            f.write('</head>\n')
            f.write('<body>\n')

            sent_id = 0

            for row in self.ws.iter_rows(min_row=2, values_only=True):
                if row and all(row):

                    sent_id += 1
                    para = ET.Element("para")
                    para.set('id', str(sent_id))

                    dct = dict(zip(self.header, row))

                    for l in self.header:
                        self.insert_into_tree(para_=para, sent_=dct[l], lang=self.LANG_COLUMNS[l])

                    f.write(ET.tostring(para, method='xml', encoding='unicode')+'\n')

            f.write('</body>\n')
            f.write('</root>\n')

            self.stats = (self.FILEPATH, str(sent_id), str(self.am_word_count), str(self.ru_word_count))

        # if move_processed:
        #     os.replace(self.FILEPATH, os.path.join(self.OUTPUT_PATH, self.FILENAME))
